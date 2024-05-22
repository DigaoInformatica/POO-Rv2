import openpyxl

# Criar uma planilha 
trabalho = openpyxl.Workbook()

# Remover página 'Sheet' padrão 
default_sheet = trabalho['Sheet']
trabalho.remove(default_sheet)

# Criar abas 'LIVROS', 'USUARIOS' e 'RESERVAS'
trabalho.create_sheet('LIVROS')
trabalho.create_sheet('USUARIOS')
trabalho.create_sheet('RESERVAS')

# Salvar o arquivo na planilha 'POO.xlsx'
trabalho.save('POO.xlsx')