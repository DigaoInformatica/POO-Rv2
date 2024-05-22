import openpyxl

# Função para encontrar a próxima linha vazia em uma planilha
def encontrar_proxima_linha(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value is None:
                return cell.row
    return sheet.max_row + 1

# Classe Livro
class Livro:
    def __init__(self, titulo, autor, genero, tipo, status):
        self.titulo = titulo
        self.autor = autor
        self.genero = genero
        self.tipo = tipo
        self.status = status

    def salvar(self, workbook):
        sheet = workbook['LIVROS']
        row = encontrar_proxima_linha(sheet)
        sheet.cell(row=row, column=1, value=self.titulo)
        sheet.cell(row=row, column=2, value=self.autor)
        sheet.cell(row=row, column=3, value=self.genero)
        sheet.cell(row=row, column=4, value=self.tipo)
        sheet.cell(row=row, column=5, value=self.status)
        workbook.save('POO.xlsx')

# Classe Usuario
class Usuario:
    def __init__(self, id, nome, cpf, senha, tipo):
        self.id = id
        self.nome = nome
        self.cpf = cpf
        self.senha = senha
        self.tipo = tipo

    def salvar(self, workbook):
        sheet = workbook['USUARIOS']
        row = encontrar_proxima_linha(sheet)
        sheet.cell(row=row, column=1, value=self.id)
        sheet.cell(row=row, column=2, value=self.nome)
        sheet.cell(row=row, column=3, value=self.cpf)
        sheet.cell(row=row, column=4, value=self.senha)
        sheet.cell(row=row, column=5, value=self.tipo)
        workbook.save('POO.xlsx')

# Classe Reserva
class Reserva:
    def __init__(self, titulo, tipo, id_usuario, data_reserva, data_devolucao):
        self.titulo = titulo
        self.tipo = tipo
        self.id_usuario = id_usuario
        self.data_reserva = data_reserva
        self.data_devolucao = data_devolucao

    def salvar(self, workbook):
        sheet = workbook['RESERVAS']
        row = encontrar_proxima_linha(sheet)
        sheet.cell(row=row, column=1, value=self.titulo)
        sheet.cell(row=row, column=2, value=self.tipo)
        sheet.cell(row=row, column=3, value=self.id_usuario)
        sheet.cell(row=row, column=4, value=self.data_reserva)
        sheet.cell(row=row, column=5, value=self.data_devolucao)
        workbook.save('POO.xlsx')

# Classe Sistema
class Sistema:
    def __init__(self, arquivo_excel):
        self.arquivo_excel = arquivo_excel
        self.workbook = openpyxl.load_workbook(arquivo_excel)

    def adicionar_livro(self, livro):
        livro.salvar(self.workbook)

    def adicionar_usuario(self, usuario):
        usuario.salvar(self.workbook)

    def adicionar_reserva(self, reserva):
        reserva.salvar(self.workbook)

    def listar(self, tipo):
        sheet = self.workbook[tipo.upper()]
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=5, values_only=True):
            if any(cell is not None for cell in row):
                print(row)

    def excluir(self, tipo, identificador):
        sheet = self.workbook[tipo.upper()]
        col_index = 1  # Identificadores estão na coluna 1 (A)
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=5):
            if row[col_index-1].value == identificador:
                sheet.delete_rows(row[0].row)
                break
        self.workbook.save(self.arquivo_excel)

    def salvar(self):
        self.workbook.save(self.arquivo_excel)